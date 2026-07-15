import re
import sqlite3
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DB_PATH = Path("data/db/nifty100.db")
CONFIG_PATH = Path("config/screener_config.yaml")
OUTPUT_PATH = Path("output")

OUTPUT_PATH.mkdir(parents=True, exist_ok=True)


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as file:
        config = yaml.safe_load(file)

    if not config or "presets" not in config:
        raise ValueError(
            "config/screener_config.yaml me presets section nahi mila."
        )

    return config


def normalize_year(value):
    if value is None or pd.isna(value):
        return None

    text = str(value).strip()

    if text.upper() == "TTM":
        return "TTM"

    if re.fullmatch(r"\d{4}-\d{2}", text):
        return text

    if re.fullmatch(r"\d{4}", text):
        return f"{text}-03"

    parsed = pd.to_datetime(
        text,
        errors="coerce"
    )

    if pd.isna(parsed):
        return text

    return parsed.strftime("%Y-%m")


def year_sort_value(value):
    if value is None or pd.isna(value):
        return -1

    text = str(value).strip()

    if text.upper() == "TTM":
        return 999999

    match = re.fullmatch(
        r"(\d{4})-(\d{2})",
        text
    )

    if match:
        year = int(match.group(1))
        month = int(match.group(2))

        return year * 100 + month

    return -1


def load_screener_data():
    connection = sqlite3.connect(DB_PATH)

    try:
        ratios = pd.read_sql_query(
            "SELECT * FROM financial_ratios",
            connection
        )

        market_cap = pd.read_sql_query(
            "SELECT * FROM market_cap",
            connection
        )

        profit_loss = pd.read_sql_query(
            """
            SELECT
                company_id,
                year,
                sales,
                net_profit
            FROM profitandloss
            """,
            connection
        )

        companies = pd.read_sql_query(
            """
            SELECT
                id,
                company_name
            FROM companies
            """,
            connection
        )

    finally:
        connection.close()

    ratios["year"] = ratios["year"].apply(
        normalize_year
    )

    market_cap["year"] = market_cap["year"].apply(
        normalize_year
    )

    profit_loss["year"] = profit_loss["year"].apply(
        normalize_year
    )

    market_cap = market_cap.drop_duplicates(
        subset=["company_id", "year"],
        keep="last"
    )

    profit_loss = profit_loss.drop_duplicates(
        subset=["company_id", "year"],
        keep="last"
    )

    companies = companies.rename(
        columns={
            "id": "company_id"
        }
    )

    dataframe = ratios.merge(
        market_cap[
            [
                "company_id",
                "year",
                "market_cap_crore",
                "enterprise_value_crore",
                "pe_ratio",
                "pb_ratio",
                "ev_ebitda",
                "dividend_yield_pct"
            ]
        ],
        on=["company_id", "year"],
        how="left"
    )

    dataframe = dataframe.merge(
        profit_loss,
        on=["company_id", "year"],
        how="left"
    )

    dataframe = dataframe.merge(
        companies[
            [
                "company_id",
                "company_name"
            ]
        ],
        on="company_id",
        how="left"
    )

    dataframe["_year_sort"] = dataframe["year"].apply(
        year_sort_value
    )

    dataframe = dataframe.sort_values(
        [
            "company_id",
            "_year_sort"
        ]
    )

    return dataframe


def latest_non_null(group, column):
    if column not in group.columns:
        return None

    values = group[column].dropna()

    if values.empty:
        return None

    return values.iloc[-1]


def build_company_snapshot(dataframe):
    rows = []

    required_columns = [
        "return_on_equity_pct",
        "return_on_capital_employed_pct",
        "return_on_assets_pct",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "debt_to_equity",
        "interest_coverage",
        "icr_label",
        "free_cash_flow_cr",
        "cfo_quality_score",
        "cfo_quality_label",
        "revenue_cagr_3yr",
        "revenue_cagr_5yr",
        "revenue_cagr_10yr",
        "pat_cagr_3yr",
        "pat_cagr_5yr",
        "pat_cagr_10yr",
        "eps_cagr_3yr",
        "eps_cagr_5yr",
        "eps_cagr_10yr",
        "asset_turnover",
        "dividend_payout_ratio_pct",
        "sales",
        "net_profit",
        "earnings_per_share",
        "market_cap_crore",
        "enterprise_value_crore",
        "pe_ratio",
        "pb_ratio",
        "ev_ebitda",
        "dividend_yield_pct",
        "composite_quality_score"
    ]

    for company_id, group in dataframe.groupby(
        "company_id"
    ):
        group = group.sort_values(
            "_year_sort"
        ).copy()

        snapshot = {
            "company_id": company_id,
            "company_name": latest_non_null(
                group,
                "company_name"
            ),
            "year": latest_non_null(
                group,
                "year"
            ),
            "broad_sector": latest_non_null(
                group,
                "broad_sector"
            )
        }

        for column in required_columns:
            snapshot[column] = latest_non_null(
                group,
                column
            )

        annual_group = group[
            group["year"]
            .astype(str)
            .str.match(
                r"^\d{4}-\d{2}$",
                na=False
            )
        ].copy()

        debt_history = pd.to_numeric(
            annual_group["debt_to_equity"],
            errors="coerce"
        ).dropna()

        if len(debt_history) >= 2:
            snapshot["de_previous"] = (
                debt_history.iloc[-2]
            )

            snapshot["de_declining_flag"] = bool(
                debt_history.iloc[-1]
                < debt_history.iloc[-2]
            )

        else:
            snapshot["de_previous"] = None
            snapshot["de_declining_flag"] = False

        if snapshot.get("icr_label") == "Debt Free":
            snapshot["icr_numeric"] = float("inf")
        else:
            snapshot["icr_numeric"] = snapshot.get(
                "interest_coverage"
            )

        rows.append(snapshot)

    return pd.DataFrame(rows)


def winsorized_score(
    series,
    reverse=False
):
    numeric = pd.to_numeric(
        series,
        errors="coerce"
    )

    valid = numeric.dropna()

    if valid.empty:
        return pd.Series(
            50.0,
            index=series.index
        )

    lower = valid.quantile(0.10)
    upper = valid.quantile(0.90)

    clipped = numeric.clip(
        lower=lower,
        upper=upper
    )

    minimum = clipped.min()
    maximum = clipped.max()

    if (
        pd.isna(minimum)
        or pd.isna(maximum)
        or minimum == maximum
    ):
        score = pd.Series(
            50.0,
            index=series.index
        )

    else:
        score = (
            (clipped - minimum)
            / (maximum - minimum)
            * 100
        )

    if reverse:
        score = 100 - score

    return score.fillna(50).clip(
        lower=0,
        upper=100
    )


def add_composite_scores(dataframe):
    dataframe = dataframe.copy()

    dataframe["score_roe"] = winsorized_score(
        dataframe["return_on_equity_pct"]
    )

    dataframe["score_roce"] = winsorized_score(
        dataframe["return_on_capital_employed_pct"]
    )

    dataframe["score_npm"] = winsorized_score(
        dataframe["net_profit_margin_pct"]
    )

    dataframe["score_fcf"] = winsorized_score(
        dataframe["free_cash_flow_cr"]
    )

    dataframe["score_cfo"] = winsorized_score(
        dataframe["cfo_quality_score"]
    )

    dataframe["score_fcf_positive"] = (
        pd.to_numeric(
            dataframe["free_cash_flow_cr"],
            errors="coerce"
        )
        .gt(0)
        .astype(int)
        * 100
    )

    dataframe["score_revenue_growth"] = (
        winsorized_score(
            dataframe["revenue_cagr_5yr"]
        )
    )

    dataframe["score_pat_growth"] = (
        winsorized_score(
            dataframe["pat_cagr_5yr"]
        )
    )

    dataframe["score_de"] = winsorized_score(
        dataframe["debt_to_equity"],
        reverse=True
    )

    finite_icr = pd.to_numeric(
        dataframe["icr_numeric"],
        errors="coerce"
    ).replace(
        [
            float("inf"),
            -float("inf")
        ],
        pd.NA
    )

    dataframe["score_icr"] = winsorized_score(
        finite_icr
    )

    dataframe["sprint3_composite_score"] = (
        dataframe["score_roe"] * 0.15
        + dataframe["score_roce"] * 0.10
        + dataframe["score_npm"] * 0.10
        + dataframe["score_fcf"] * 0.15
        + dataframe["score_cfo"] * 0.10
        + dataframe["score_fcf_positive"] * 0.05
        + dataframe["score_revenue_growth"] * 0.10
        + dataframe["score_pat_growth"] * 0.10
        + dataframe["score_de"] * 0.10
        + dataframe["score_icr"] * 0.05
    ).round(2)

    dataframe["sector_relative_score"] = (
        dataframe.groupby(
            "broad_sector"
        )[
            "sprint3_composite_score"
        ]
        .transform(
            winsorized_score
        )
        .round(2)
    )

    return dataframe


def numeric_column(
    dataframe,
    column
):
    return pd.to_numeric(
        dataframe[column],
        errors="coerce"
    )


def apply_filters(
    dataframe,
    rules
):
    result = dataframe.copy()

    if "roe_min" in rules:
        result = result[
            numeric_column(
                result,
                "return_on_equity_pct"
            ) > rules["roe_min"]
        ]

    if "de_max" in rules:
        debt = numeric_column(
            result,
            "debt_to_equity"
        )

        result = result[
            result["broad_sector"].eq(
                "Financials"
            )
            | debt.lt(
                rules["de_max"]
            )
        ]

    if "de_equals" in rules:
        result = result[
            numeric_column(
                result,
                "debt_to_equity"
            ).eq(
                rules["de_equals"]
            )
        ]

    if "de_practical_max" in rules:
        result = result[
            numeric_column(
                result,
                "debt_to_equity"
            ).le(
                rules["de_practical_max"]
            )
        ]

    if "fcf_min" in rules:
        result = result[
            numeric_column(
                result,
                "free_cash_flow_cr"
            ) > rules["fcf_min"]
        ]

    if "revenue_cagr_5yr_min" in rules:
        result = result[
            numeric_column(
                result,
                "revenue_cagr_5yr"
            ) > rules[
                "revenue_cagr_5yr_min"
            ]
        ]

    if "revenue_cagr_3yr_min" in rules:
        result = result[
            numeric_column(
                result,
                "revenue_cagr_3yr"
            ) > rules[
                "revenue_cagr_3yr_min"
            ]
        ]

    if "pat_cagr_5yr_min" in rules:
        result = result[
            numeric_column(
                result,
                "pat_cagr_5yr"
            ) > rules[
                "pat_cagr_5yr_min"
            ]
        ]

    if "opm_min" in rules:
        result = result[
            numeric_column(
                result,
                "operating_profit_margin_pct"
            ) > rules["opm_min"]
        ]

    if "pe_max" in rules:
        result = result[
            numeric_column(
                result,
                "pe_ratio"
            ) < rules["pe_max"]
        ]

    if "pb_max" in rules:
        result = result[
            numeric_column(
                result,
                "pb_ratio"
            ) < rules["pb_max"]
        ]

    if "dividend_yield_min" in rules:
        result = result[
            numeric_column(
                result,
                "dividend_yield_pct"
            ) > rules[
                "dividend_yield_min"
            ]
        ]

    if "dividend_payout_max" in rules:
        result = result[
            numeric_column(
                result,
                "dividend_payout_ratio_pct"
            ) < rules[
                "dividend_payout_max"
            ]
        ]

    if "icr_min" in rules:
        result = result[
            pd.to_numeric(
                result["icr_numeric"],
                errors="coerce"
            ) > rules["icr_min"]
        ]

    if "market_cap_min" in rules:
        result = result[
            numeric_column(
                result,
                "market_cap_crore"
            ) > rules[
                "market_cap_min"
            ]
        ]

    if "net_profit_min" in rules:
        result = result[
            numeric_column(
                result,
                "net_profit"
            ) > rules[
                "net_profit_min"
            ]
        ]

    if "eps_cagr_min" in rules:
        result = result[
            numeric_column(
                result,
                "eps_cagr_5yr"
            ) > rules[
                "eps_cagr_min"
            ]
        ]

    if "asset_turnover_min" in rules:
        result = result[
            numeric_column(
                result,
                "asset_turnover"
            ) > rules[
                "asset_turnover_min"
            ]
        ]

    if "sales_min" in rules:
        result = result[
            numeric_column(
                result,
                "sales"
            ) > rules[
                "sales_min"
            ]
        ]

    if rules.get(
        "de_declining",
        False
    ):
        result = result[
            result[
                "de_declining_flag"
            ].eq(True)
        ]

    return result.sort_values(
        "sprint3_composite_score",
        ascending=False,
        na_position="last"
    )


def safe_number(value):
    if value is None or pd.isna(value):
        return None

    try:
        return float(value)

    except (TypeError, ValueError):
        return None


def format_excel(
    file_path,
    sheet_rules
):
    workbook = load_workbook(
        file_path
    )

    green = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    red = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    rule_column_map = {
        "roe_min": (
            "return_on_equity_pct",
            "min"
        ),
        "de_max": (
            "debt_to_equity",
            "max"
        ),
        "de_equals": (
            "debt_to_equity",
            "equal"
        ),
        "de_practical_max": (
            "debt_to_equity",
            "max_equal"
        ),
        "fcf_min": (
            "free_cash_flow_cr",
            "min"
        ),
        "revenue_cagr_5yr_min": (
            "revenue_cagr_5yr",
            "min"
        ),
        "revenue_cagr_3yr_min": (
            "revenue_cagr_3yr",
            "min"
        ),
        "pat_cagr_5yr_min": (
            "pat_cagr_5yr",
            "min"
        ),
        "pe_max": (
            "pe_ratio",
            "max"
        ),
        "pb_max": (
            "pb_ratio",
            "max"
        ),
        "dividend_yield_min": (
            "dividend_yield_pct",
            "min"
        ),
        "dividend_payout_max": (
            "dividend_payout_ratio_pct",
            "max"
        ),
        "sales_min": (
            "sales",
            "min"
        )
    }

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[
            sheet_name
        ]

        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        rules = sheet_rules.get(
            sheet_name,
            {}
        )

        for (
            rule_name,
            threshold
        ) in rules.items():

            if rule_name not in rule_column_map:
                continue

            (
                column_name,
                comparison
            ) = rule_column_map[
                rule_name
            ]

            if column_name not in headers:
                continue

            column_number = headers[
                column_name
            ]

            for row_number in range(
                2,
                worksheet.max_row + 1
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column_number
                )

                numeric_value = safe_number(
                    cell.value
                )

                if numeric_value is None:
                    cell.fill = red
                    continue

                if comparison == "min":
                    passed = (
                        numeric_value
                        > threshold
                    )

                elif comparison == "max":
                    passed = (
                        numeric_value
                        < threshold
                    )

                elif comparison == "max_equal":
                    passed = (
                        numeric_value
                        <= threshold
                    )

                else:
                    passed = (
                        numeric_value
                        == threshold
                    )

                cell.fill = (
                    green
                    if passed
                    else red
                )

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(
                column_cells[0].column
            )

            maximum_length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                maximum_length + 2,
                28
            )

    workbook.save(
        file_path
    )


def run_screener_engine():
    config = load_config()

    historical_data = (
        load_screener_data()
    )

    company_snapshot = (
        build_company_snapshot(
            historical_data
        )
    )

    company_snapshot = (
        add_composite_scores(
            company_snapshot
        )
    )

    company_snapshot.to_csv(
        OUTPUT_PATH
        / "screener_company_snapshot.csv",
        index=False
    )

    output_file = (
        OUTPUT_PATH
        / "screener_output.xlsx"
    )

    selected_columns = [
        "company_id",
        "company_name",
        "year",
        "broad_sector",
        "return_on_equity_pct",
        "return_on_capital_employed_pct",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "debt_to_equity",
        "interest_coverage",
        "icr_label",
        "free_cash_flow_cr",
        "revenue_cagr_3yr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "sales",
        "net_profit",
        "pe_ratio",
        "pb_ratio",
        "dividend_yield_pct",
        "dividend_payout_ratio_pct",
        "market_cap_crore",
        "sprint3_composite_score",
        "sector_relative_score",
        "de_declining_flag"
    ]

    selected_columns = [
        column
        for column in selected_columns
        if column
        in company_snapshot.columns
    ]

    preset_counts = {}
    sheet_rules = {}

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl"
    ) as writer:

        for (
            preset_name,
            rules
        ) in config[
            "presets"
        ].items():

            result = apply_filters(
                company_snapshot,
                rules
            )

            if len(result) > 50:
                result = result.head(
                    50
                )

            preset_counts[
                preset_name
            ] = len(result)

            sheet_name = (
                preset_name[:31]
            )

            sheet_rules[
                sheet_name
            ] = rules

            result[
                selected_columns
            ].to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

            result[
                selected_columns
            ].to_csv(
                OUTPUT_PATH
                / f"{preset_name}.csv",
                index=False
            )

    format_excel(
        output_file,
        sheet_rules
    )

    print(
        "output/screener_output.xlsx generated"
    )

    print(
        "output/screener_company_snapshot.csv generated"
    )

    print(
        "Selection Mode: Latest non-null KPI snapshot including TTM"
    )

    print()

    print(
        "Snapshot Companies:",
        company_snapshot[
            "company_id"
        ].nunique()
    )

    print(
        "Revenue CAGR 5Y Available:",
        company_snapshot[
            "revenue_cagr_5yr"
        ].notna().sum()
    )

    print(
        "PAT CAGR 5Y Available:",
        company_snapshot[
            "pat_cagr_5yr"
        ].notna().sum()
    )

    print()

    for (
        name,
        count
    ) in preset_counts.items():
        print(
            f"{name}: {count}"
        )

    all_pass = all(
        5 <= count <= 50
        for count
        in preset_counts.values()
    )

    print()

    print(
        "All presets count pass:",
        all_pass
    )


if __name__ == "__main__":
    run_screener_engine()