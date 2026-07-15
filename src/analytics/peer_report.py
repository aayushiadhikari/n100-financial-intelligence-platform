import re
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


DB_PATH = Path("data/db/nifty100.db")
OUTPUT_PATH = Path("output")
OUTPUT_FILE = OUTPUT_PATH / "peer_comparison.xlsx"

OUTPUT_PATH.mkdir(parents=True, exist_ok=True)


BASE_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "capex_cr",
    "earnings_per_share",
    "dividend_payout_ratio_pct",
    "total_debt_cr",
    "cash_from_operations_cr",
    "revenue_cagr_3yr",
    "revenue_cagr_5yr",
    "pat_cagr_3yr",
    "pat_cagr_5yr",
    "eps_cagr_3yr",
    "eps_cagr_5yr",
    "composite_quality_score"
]


PERCENTILE_METRICS = {
    "roe": "roe_percentile",
    "roce": "roce_percentile",
    "npm": "npm_percentile",
    "debt_to_equity": "debt_to_equity_percentile",
    "fcf": "fcf_percentile",
    "pat_cagr_5yr": "pat_cagr_5yr_percentile",
    "revenue_cagr_5yr": "revenue_cagr_5yr_percentile",
    "eps_cagr_5yr": "eps_cagr_5yr_percentile",
    "interest_coverage": "interest_coverage_percentile",
    "asset_turnover": "asset_turnover_percentile"
}


def load_data():
    connection = sqlite3.connect(DB_PATH)

    ratios = pd.read_sql_query(
        "SELECT * FROM financial_ratios",
        connection
    )

    peer_groups = pd.read_sql_query(
        "SELECT * FROM peer_groups",
        connection
    )

    companies = pd.read_sql_query(
        "SELECT id, company_name FROM companies",
        connection
    )

    peer_percentiles = pd.read_sql_query(
        "SELECT * FROM peer_percentiles",
        connection
    )

    connection.close()

    return (
        ratios,
        peer_groups,
        companies,
        peer_percentiles
    )


def get_latest_annual_data(ratios):
    annual = ratios[
        ratios["year"]
        .astype(str)
        .str.match(r"^\d{4}-\d{2}$", na=False)
    ].copy()

    if annual.empty:
        raise ValueError(
            "financial_ratios table me annual data nahi mila."
        )

    latest_year = annual["year"].max()

    latest = annual[
        annual["year"] == latest_year
    ].copy()

    return latest, latest_year


def prepare_percentile_wide(peer_percentiles, latest_year):
    latest_percentiles = peer_percentiles[
        peer_percentiles["year"] == latest_year
    ].copy()

    latest_percentiles["percentile_column"] = (
        latest_percentiles["metric"]
        .map(PERCENTILE_METRICS)
    )

    latest_percentiles = latest_percentiles[
        latest_percentiles["percentile_column"].notna()
    ].copy()

    wide = latest_percentiles.pivot_table(
        index=[
            "company_id",
            "peer_group_name"
        ],
        columns="percentile_column",
        values="percentile_rank",
        aggfunc="first"
    ).reset_index()

    wide.columns.name = None

    return wide


def clean_sheet_name(name):
    cleaned = re.sub(
        r'[\[\]:*?/\\]',
        "_",
        str(name)
    )

    cleaned = cleaned.strip()

    if not cleaned:
        cleaned = "Peer_Group"

    return cleaned[:31]


def create_unique_sheet_name(name, used_names):
    base = clean_sheet_name(name)
    candidate = base
    counter = 1

    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = (
            base[:31 - len(suffix)]
            + suffix
        )
        counter += 1

    used_names.add(candidate)

    return candidate


def build_peer_dataset():
    (
        ratios,
        peer_groups,
        companies,
        peer_percentiles
    ) = load_data()

    latest, latest_year = get_latest_annual_data(
        ratios
    )

    percentile_wide = prepare_percentile_wide(
        peer_percentiles,
        latest_year
    )

    company_columns = companies[
        [
            "id",
            "company_name"
        ]
    ].rename(
        columns={
            "id": "company_id"
        }
    )

    peer_data = peer_groups.merge(
        company_columns,
        on="company_id",
        how="left"
    )

    peer_data = peer_data.merge(
        latest,
        on="company_id",
        how="left"
    )

    peer_data = peer_data.merge(
        percentile_wide,
        on=[
            "company_id",
            "peer_group_name"
        ],
        how="left"
    )

    return peer_data, latest_year


def selected_output_columns(peer_data):
    columns = [
        "company_id",
        "company_name",
        "peer_group_name",
        "is_benchmark",
        "year"
    ]

    columns.extend(
        [
            column
            for column in BASE_METRICS
            if column in peer_data.columns
        ]
    )

    columns.extend(
        [
            percentile_column
            for percentile_column
            in PERCENTILE_METRICS.values()
            if percentile_column in peer_data.columns
        ]
    )

    return columns


def add_median_row(group_df):
    median_row = {}

    for column in group_df.columns:
        if column == "company_id":
            median_row[column] = "PEER_MEDIAN"

        elif column == "company_name":
            median_row[column] = "Peer Group Median"

        elif column == "peer_group_name":
            median_row[column] = (
                group_df["peer_group_name"].iloc[0]
            )

        elif column == "is_benchmark":
            median_row[column] = False

        elif column == "year":
            median_row[column] = group_df["year"].iloc[0]

        elif pd.api.types.is_numeric_dtype(
            group_df[column]
        ):
            median_row[column] = (
                pd.to_numeric(
                    group_df[column],
                    errors="coerce"
                ).median()
            )

        else:
            median_row[column] = None

    median_df = pd.DataFrame(
        [median_row]
    )

    return pd.concat(
        [
            group_df,
            median_df
        ],
        ignore_index=True
    )


def write_excel(peer_data):
    output_columns = selected_output_columns(
        peer_data
    )

    used_sheet_names = set()
    sheet_mapping = {}

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        for peer_group_name, group in peer_data.groupby(
            "peer_group_name",
            sort=True
        ):
            group = group.copy()

            group = group.sort_values(
                "composite_quality_score",
                ascending=False,
                na_position="last"
            )

            group = group[output_columns]

            group = add_median_row(
                group
            )

            sheet_name = create_unique_sheet_name(
                peer_group_name,
                used_sheet_names
            )

            sheet_mapping[sheet_name] = peer_group_name

            group.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

    return sheet_mapping


def apply_excel_formatting(sheet_mapping):
    workbook = load_workbook(
        OUTPUT_FILE
    )

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    yellow_fill = PatternFill(
        start_color="FFF2CC",
        end_color="FFF2CC",
        fill_type="solid"
    )

    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    benchmark_fill = PatternFill(
        start_color="FFD966",
        end_color="FFD966",
        fill_type="solid"
    )

    median_fill = PatternFill(
        start_color="D9EAD3",
        end_color="D9EAD3",
        fill_type="solid"
    )

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        benchmark_column = headers.get(
            "is_benchmark"
        )

        company_id_column = headers.get(
            "company_id"
        )

        percentile_columns = [
            column_index
            for header, column_index in headers.items()
            if isinstance(header, str)
            and header.endswith("_percentile")
        ]

        for row_number in range(
            2,
            worksheet.max_row + 1
        ):
            company_id = (
                worksheet.cell(
                    row=row_number,
                    column=company_id_column
                ).value
                if company_id_column
                else None
            )

            is_median_row = (
                company_id == "PEER_MEDIAN"
            )

            is_benchmark = False

            if benchmark_column:
                benchmark_value = worksheet.cell(
                    row=row_number,
                    column=benchmark_column
                ).value

                is_benchmark = benchmark_value in (
                    True,
                    1,
                    "1",
                    "True",
                    "TRUE"
                )

            if is_median_row:
                for cell in worksheet[row_number]:
                    cell.fill = median_fill
                    cell.font = Font(
                        bold=True
                    )

            elif is_benchmark:
                for cell in worksheet[row_number]:
                    cell.fill = benchmark_fill
                    cell.font = Font(
                        bold=True
                    )

            for column_index in percentile_columns:
                cell = worksheet.cell(
                    row=row_number,
                    column=column_index
                )

                value = cell.value

                if (
                    is_median_row
                    or value is None
                ):
                    continue

                try:
                    numeric_value = float(value)
                except (TypeError, ValueError):
                    continue

                if numeric_value >= 0.75:
                    cell.fill = green_fill

                elif numeric_value <= 0.25:
                    cell.fill = red_fill

                else:
                    cell.fill = yellow_fill

                cell.number_format = "0.00%"

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(
                column_cells[0].column
            )

            maximum_length = 0

            for cell in column_cells:
                value = (
                    ""
                    if cell.value is None
                    else str(cell.value)
                )

                maximum_length = max(
                    maximum_length,
                    len(value)
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                maximum_length + 2,
                28
            )

        worksheet.row_dimensions[1].height = 32

    workbook.save(
        OUTPUT_FILE
    )


def validate_excel(expected_peer_groups):
    workbook = load_workbook(
        OUTPUT_FILE,
        read_only=True
    )

    actual_sheets = workbook.sheetnames

    workbook.close()

    print(
        "Peer Comparison Sheets:",
        len(actual_sheets)
    )

    print(
        "Expected Peer Groups:",
        expected_peer_groups
    )

    print(
        "Exactly 11 Sheets:",
        len(actual_sheets) == 11
    )


def generate_peer_comparison_report():
    peer_data, latest_year = build_peer_dataset()

    peer_group_count = (
        peer_data["peer_group_name"]
        .nunique()
    )

    sheet_mapping = write_excel(
        peer_data
    )

    apply_excel_formatting(
        sheet_mapping
    )

    validate_excel(
        peer_group_count
    )

    print()
    print(
        "output/peer_comparison.xlsx generated successfully."
    )

    print(
        "Latest Year:",
        latest_year
    )

    print(
        "Peer Groups:",
        peer_group_count
    )

    print(
        "Companies:",
        peer_data["company_id"].nunique()
    )


if __name__ == "__main__":
    generate_peer_comparison_report()