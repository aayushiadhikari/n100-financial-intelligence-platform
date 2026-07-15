import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


DB_PATH = Path("data/db/nifty100.db")
OUTPUT_PATH = Path("output")
RADAR_PATH = Path("reports/radar_charts")


def main():
    connection = sqlite3.connect(DB_PATH)

    peer_percentiles = pd.read_sql_query(
        "SELECT * FROM peer_percentiles",
        connection
    )

    connection.close()

    screener_file = OUTPUT_PATH / "screener_output.xlsx"
    peer_file = OUTPUT_PATH / "peer_comparison.xlsx"

    screener_workbook = load_workbook(
        screener_file,
        read_only=True
    )

    peer_workbook = load_workbook(
        peer_file,
        read_only=True
    )

    screener_sheet_count = len(
        screener_workbook.sheetnames
    )

    peer_sheet_count = len(
        peer_workbook.sheetnames
    )

    screener_workbook.close()
    peer_workbook.close()

    screener_counts = {}

    excel_file = pd.ExcelFile(
        screener_file
    )

    for sheet_name in excel_file.sheet_names:
        dataframe = pd.read_excel(
            excel_file,
            sheet_name=sheet_name
        )

        screener_counts[sheet_name] = len(
            dataframe
        )

    all_screeners_pass = all(
        5 <= count <= 50
        for count in screener_counts.values()
    )

    radar_count = len(
        list(
            RADAR_PATH.glob("*.png")
        )
    )

    summary = {
        "screener_output_exists": screener_file.exists(),
        "screener_sheet_count": screener_sheet_count,
        "exactly_6_screener_sheets": screener_sheet_count == 6,
        "all_screeners_between_5_and_50": all_screeners_pass,
        "peer_comparison_exists": peer_file.exists(),
        "peer_sheet_count": peer_sheet_count,
        "exactly_11_peer_sheets": peer_sheet_count == 11,
        "peer_percentile_rows": len(peer_percentiles),
        "peer_groups": peer_percentiles[
            "peer_group_name"
        ].nunique(),
        "peer_groups_pass": peer_percentiles[
            "peer_group_name"
        ].nunique() == 11,
        "radar_chart_count": radar_count,
        "radar_charts_pass": radar_count >= 56,
        "dq_tests_expected": 14
    }

    summary_df = pd.DataFrame(
        [summary]
    )

    summary_df.to_csv(
        OUTPUT_PATH / "sprint3_review_summary.csv",
        index=False
    )

    counts_df = pd.DataFrame(
        [
            {
                "preset": preset,
                "company_count": count,
                "count_pass": 5 <= count <= 50
            }
            for preset, count in screener_counts.items()
        ]
    )

    counts_df.to_csv(
        OUTPUT_PATH / "sprint3_preset_counts.csv",
        index=False
    )

    print("Sprint 3 Review Completed")
    print()
    print(summary_df.T)
    print()
    print("Preset Counts")
    print(counts_df.to_string(index=False))


if __name__ == "__main__":
    main()