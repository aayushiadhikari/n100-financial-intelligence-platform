import sqlite3
import sys
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
sys.path.append(str(ROOT))

DB_PATH = Path("data/db/nifty100.db")
OUTPUT_PATH = Path("output")
CONFIG_PATH = Path("config/screener_config.yaml")
RADAR_PATH = Path("reports/radar_charts")


def read_table(table_name):
    connection = sqlite3.connect(DB_PATH)

    dataframe = pd.read_sql_query(
        f"SELECT * FROM {table_name}",
        connection
    )

    connection.close()

    return dataframe


def read_config():
    with open(
        CONFIG_PATH,
        "r",
        encoding="utf-8"
    ) as file:
        return yaml.safe_load(file)


def test_01_database_exists():
    assert DB_PATH.exists()


def test_02_financial_ratios_table_has_rows():
    ratios = read_table("financial_ratios")

    assert len(ratios) >= 1100


def test_03_screener_config_exists():
    assert CONFIG_PATH.exists()


def test_04_exactly_six_presets_exist():
    config = read_config()

    assert "presets" in config
    assert len(config["presets"]) == 6


def test_05_required_presets_exist():
    config = read_config()

    expected = {
        "quality_compounder",
        "value_pick",
        "growth_accelerator",
        "dividend_champion",
        "debt_free_blue_chip",
        "turnaround_watch"
    }

    assert set(config["presets"].keys()) == expected


def test_06_screener_output_exists():
    assert (
        OUTPUT_PATH /
        "screener_output.xlsx"
    ).exists()


def test_07_screener_output_has_six_sheets():
    workbook = load_workbook(
        OUTPUT_PATH / "screener_output.xlsx",
        read_only=True
    )

    assert len(workbook.sheetnames) == 6

    workbook.close()


def test_08_each_screener_has_between_5_and_50_rows():
    workbook = pd.ExcelFile(
        OUTPUT_PATH / "screener_output.xlsx"
    )

    for sheet_name in workbook.sheet_names:
        dataframe = pd.read_excel(
            workbook,
            sheet_name=sheet_name
        )

        assert 5 <= len(dataframe) <= 50, (
            f"{sheet_name} returned {len(dataframe)} rows"
        )


def test_09_quality_compounder_core_rules():
    dataframe = pd.read_excel(
        OUTPUT_PATH / "screener_output.xlsx",
        sheet_name="quality_compounder"
    )

    assert (
        dataframe["return_on_equity_pct"] > 15
    ).all()

    non_financial = dataframe[
        dataframe["broad_sector"] != "Financials"
    ]

    assert (
        non_financial["debt_to_equity"] < 1
    ).all()

    assert (
        dataframe["free_cash_flow_cr"] > 0
    ).all()

    assert (
        dataframe["revenue_cagr_5yr"] > 10
    ).all()


def test_10_peer_percentiles_table_exists_and_has_rows():
    percentiles = read_table(
        "peer_percentiles"
    )

    assert len(percentiles) > 0


def test_11_peer_percentiles_cover_11_groups():
    percentiles = read_table(
        "peer_percentiles"
    )

    assert (
        percentiles["peer_group_name"]
        .nunique()
        == 11
    )


def test_12_peer_percentile_values_are_valid():
    percentiles = read_table(
        "peer_percentiles"
    )

    numeric = pd.to_numeric(
        percentiles["percentile_rank"],
        errors="coerce"
    )

    assert numeric.notna().all()
    assert numeric.between(0, 1).all()


def test_13_peer_comparison_has_exactly_11_sheets():
    workbook = load_workbook(
        OUTPUT_PATH / "peer_comparison.xlsx",
        read_only=True
    )

    assert len(workbook.sheetnames) == 11

    workbook.close()


def test_14_radar_charts_generated():
    charts = list(
        RADAR_PATH.glob("*.png")
    )

    assert len(charts) >= 56